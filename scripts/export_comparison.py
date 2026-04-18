#!/usr/bin/env python3
"""
将 Google Maps API 和 Scraper 的抓取结果导出到同一个 Excel 的两个 Sheet 中，便于人工对比。

运行前准备:
    1. 确保爬虫服务已启动
    2. 设置环境变量: export GOOGLE_MAPS_API_KEY="your_api_key"
    3. 安装依赖: pip install openpyxl requests

运行方式:
    python scripts/export_comparison.py
    python scripts/export_comparison.py --city Beijing   # 指定城市
    python scripts/export_comparison.py --lat 39.9 --lng 116.4 --radius 2000 --keywords restaurant

输出文件:
    ./compare_results_20250418_153022.xlsx （自动命名）
"""

import json
import os
import random
import sys
import time
from dataclasses import dataclass
from math import atan2, cos, radians, sin, sqrt
from typing import Any, Dict, List, Optional, Tuple

import requests

# openpyxl 是可选依赖，没有安装时给出友好提示
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("错误: 请先安装 openpyxl")
    print("  pip install openpyxl")
    sys.exit(1)


# ===================== 城市配置 =====================

CITIES = [
    {"name": "Cairo", "lat": 30.0444, "lng": 31.2357, "radius": 3000, "keywords": "hospital"},
    {"name": "Beijing", "lat": 39.9042, "lng": 116.4074, "radius": 3000, "keywords": "restaurant"},
    {"name": "New York", "lat": 40.7128, "lng": -74.0060, "radius": 3000, "keywords": "coffee shop"},
    {"name": "London", "lat": 51.5074, "lng": -0.1278, "radius": 3000, "keywords": "hotel"},
    {"name": "Tokyo", "lat": 35.6762, "lng": 139.6503, "radius": 3000, "keywords": "pharmacy"},
    {"name": "Paris", "lat": 48.8566, "lng": 2.3522, "radius": 3000, "keywords": "bakery"},
    {"name": "Sydney", "lat": -33.8688, "lng": 151.2093, "radius": 3000, "keywords": "cafe"},
    {"name": "Dubai", "lat": 25.2048, "lng": 55.2708, "radius": 3000, "keywords": "mall"},
    {"name": "Mumbai", "lat": 19.0760, "lng": 72.8777, "radius": 3000, "keywords": "bank"},
    {"name": "São Paulo", "lat": -23.5505, "lng": -46.6333, "radius": 3000, "keywords": "supermarket"},
]


# ===================== 数据模型 =====================


@dataclass
class SearchConfig:
    keywords: str
    lat: float
    lng: float
    radius: int
    lang: str = "en"
    zoom: int = 14


@dataclass
class POI:
    name: str
    lat: float
    lng: float
    address: str = ""
    phone: str = ""
    website: str = ""
    category: str = ""
    rating: float = 0.0


# ===================== 坐标生成 =====================


def random_point_in_city(city: Dict[str, Any]) -> Tuple[float, float]:
    """在城市中心周围随机生成一个坐标点（均匀分布在半径范围内）"""
    center_lat = city["lat"]
    center_lng = city["lng"]
    radius_m = city["radius"]

    r = radius_m * sqrt(random.random())
    theta = random.uniform(0, 2 * 3.14159265359)

    lat_offset = r * cos(theta) / 111000.0
    lng_offset = r * sin(theta) / (111000.0 * cos(radians(center_lat)))

    return center_lat + lat_offset, center_lng + lng_offset


def pick_random_city() -> Dict[str, Any]:
    return random.choice(CITIES)


# ===================== 客户端 =====================


class GoogleMapsAPIClient:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"

    def search(self, config: SearchConfig) -> List[POI]:
        pois: List[POI] = []
        next_page_token: Optional[str] = None
        page_count = 0

        while True:
            params = {
                "key": self.api_key,
                "location": f"{config.lat},{config.lng}",
                "radius": config.radius,
                "keyword": config.keywords,
                "language": config.lang,
            }
            if next_page_token:
                params["pagetoken"] = next_page_token

            resp = requests.get(self.base_url, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()

            status = data.get("status", "")
            if status not in ("OK", "ZERO_RESULTS"):
                raise RuntimeError(
                    f"Google Maps API error: {status} — {data.get('error_message', '')}"
                )

            for result in data.get("results", []):
                loc = result.get("geometry", {}).get("location", {})
                types = result.get("types", [])
                category = types[0] if types else ""
                poi = POI(
                    name=result.get("name", ""),
                    lat=loc.get("lat", 0.0),
                    lng=loc.get("lng", 0.0),
                    address=result.get("vicinity", ""),
                    rating=result.get("rating", 0.0) or 0.0,
                    category=category,
                )
                pois.append(poi)

            next_page_token = data.get("next_page_token")
            if not next_page_token:
                break

            page_count += 1
            if page_count >= 2:
                break

            time.sleep(2)

        return pois


class ScraperClient:
    def __init__(self, base_url: str = "http://localhost:8080"):
        self.base_url = base_url.rstrip("/")

    def submit_job(self, config: SearchConfig, job_name: str = "export_test") -> str:
        geojson = {
            "type": "Feature",
            "properties": {"type": "circle", "radius": config.radius},
            "geometry": {"type": "Point", "coordinates": [config.lng, config.lat]},
        }
        payload = {
            "name": job_name,
            "keywords": [config.keywords],
            "lang": config.lang,
            "zoom": config.zoom,
            "lat": str(config.lat),
            "lon": str(config.lng),
            "geojson": json.dumps(geojson),
            "radius": config.radius,
            "fast_mode": False,
            "depth": 10,
            "max_time": 600,
            "email": False,
        }
        resp = requests.post(f"{self.base_url}/api/v1/jobs", json=payload, timeout=30)
        resp.raise_for_status()
        return resp.json()["id"]

    def get_job_status(self, job_id: str) -> Optional[str]:
        try:
            resp = requests.get(f"{self.base_url}/api/v1/jobs/{job_id}", timeout=10)
            resp.raise_for_status()
            data = resp.json()
            return data.get("status") or data.get("Status")
        except Exception:
            return None

    def get_results(self, job_id: str) -> Optional[List[POI]]:
        resp = requests.get(f"{self.base_url}/api/v1/jobs/{job_id}/pois", timeout=30)
        resp.raise_for_status()
        data = resp.json()
        pois_data = data.get("pois", []) if isinstance(data, dict) else data
        if not pois_data:
            return None
        return [
            POI(
                name=p.get("name", ""),
                lat=p.get("lat", 0.0),
                lng=p.get("lng", 0.0),
                address=p.get("address", ""),
                phone=p.get("phone", ""),
                website=p.get("website", ""),
                category=p.get("category", ""),
                rating=p.get("rating", 0.0) or 0.0,
            )
            for p in pois_data
        ]

    def search(self, config: SearchConfig, max_wait: int = 300) -> List[POI]:
        job_id = self.submit_job(config)
        print(f"[Scraper] 任务已提交: {job_id}")
        start = time.time()
        while time.time() - start < max_wait:
            status = self.get_job_status(job_id)
            if status == "ok":
                results = self.get_results(job_id)
                if results is not None:
                    return results
            elif status == "failed":
                raise RuntimeError(f"爬虫任务失败: {job_id}")
            else:
                results = self.get_results(job_id)
                if results is not None:
                    return results
            time.sleep(5)
        raise TimeoutError(f"爬虫任务 {job_id} 超时")


# ===================== Excel 导出 =====================


def create_excel(
    gm_pois: List[POI],
    sc_pois: List[POI],
    config: SearchConfig,
    city_name: str = "",
    output_dir: str = ".",
) -> str:
    """创建 Excel 文件，两个 Sheet 分别存放 API 和 Scraper 结果"""

    wb = Workbook()

    # 删除默认 Sheet，创建两个新 Sheet
    wb.remove(wb.active)
    ws_gm = wb.create_sheet(title="Google Maps API")
    ws_sc = wb.create_sheet(title="Scraper")

    # 表头
    headers = ["序号", "名称", "地址", "电话", "网站", "类别", "评分", "纬度", "经度"]

    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_sheet(ws, pois, sheet_title):
        # 写入元信息
        ws.append(["数据来源", sheet_title])
        ws.append(["关键词", config.keywords])
        ws.append(["中心纬度", config.lat])
        ws.append(["中心经度", config.lng])
        ws.append(["搜索半径(米)", config.radius])
        if city_name:
            ws.append(["城市", city_name])
        ws.append([])  # 空行

        # 写入表头
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 写入数据
        for idx, poi in enumerate(pois, start=1):
            row = [
                idx,
                poi.name,
                poi.address,
                poi.phone,
                poi.website,
                poi.category,
                poi.rating if poi.rating > 0 else "",
                poi.lat,
                poi.lng,
            ]
            ws.append(row)

        # 设置列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 35
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 14

        # 冻结首行（表头）
        ws.freeze_panes = "A9"

        # 添加边框
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    write_sheet(ws_gm, gm_pois, "Google Maps Places API")
    write_sheet(ws_sc, sc_pois, "google-maps-scraper")

    # 文件名
    from datetime import datetime

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"compare_results_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ===================== 主流程 =====================


def parse_args() -> Tuple[SearchConfig, Optional[str]]:
    """解析命令行参数，返回 (config, city_name)"""
    args = sys.argv[1:]

    # 尝试解析 --city, --lat, --lng, --radius, --keywords
    city_name = None
    lat = None
    lng = None
    radius = 3000
    keywords = None

    i = 0
    while i < len(args):
        if args[i] == "--city" and i + 1 < len(args):
            city_name = args[i + 1]
            i += 2
        elif args[i] == "--lat" and i + 1 < len(args):
            lat = float(args[i + 1])
            i += 2
        elif args[i] == "--lng" and i + 1 < len(args):
            lng = float(args[i + 1])
            i += 2
        elif args[i] == "--radius" and i + 1 < len(args):
            radius = int(args[i + 1])
            i += 2
        elif args[i] == "--keywords" and i + 1 < len(args):
            keywords = args[i + 1]
            i += 2
        else:
            i += 1

    if city_name:
        for c in CITIES:
            if c["name"].lower() == city_name.lower():
                city = c
                rand_lat, rand_lng = random_point_in_city(city)
                return (
                    SearchConfig(
                        keywords=keywords or city["keywords"],
                        lat=round(rand_lat, 6),
                        lng=round(rand_lng, 6),
                        radius=radius if radius != 3000 else city["radius"],
                        lang="en",
                        zoom=14,
                    ),
                    city["name"],
                )
        print(f"警告: 未找到城市 '{city_name}'，使用随机城市")

    if lat is not None and lng is not None:
        return (
            SearchConfig(
                keywords=keywords or "restaurant",
                lat=lat,
                lng=lng,
                radius=radius,
                lang="en",
                zoom=14,
            ),
            "自定义坐标",
        )

    # 默认随机城市
    city = pick_random_city()
    rand_lat, rand_lng = random_point_in_city(city)
    return (
        SearchConfig(
            keywords=keywords or city["keywords"],
            lat=round(rand_lat, 6),
            lng=round(rand_lng, 6),
            radius=city["radius"],
            lang="en",
            zoom=14,
        ),
        city["name"],
    )


def main():
    GOOGLE_MAPS_API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "")
    SCRAPER_BASE_URL = os.environ.get("SCRAPER_URL", "http://localhost:8080")

    if not GOOGLE_MAPS_API_KEY:
        print("错误: 请设置环境变量 GOOGLE_MAPS_API_KEY")
        print("  export GOOGLE_MAPS_API_KEY='your_api_key'")
        sys.exit(1)

    config, city_name = parse_args()

    print("=" * 65)
    print("📊 Google Maps API vs Scraper 结果导出")
    print("=" * 65)
    if city_name:
        print(f"  城市   : {city_name}")
    print(f"  关键词 : {config.keywords}")
    print(f"  坐标   : ({config.lat}, {config.lng})")
    print(f"  半径   : {config.radius} 米")

    # 1) Google Maps API
    print("\n🔎 正在调用 Google Maps Places API ...")
    gm_client = GoogleMapsAPIClient(GOOGLE_MAPS_API_KEY)
    gm_pois = gm_client.search(config)
    print(f"✅ 获取到 {len(gm_pois)} 个 POI")

    # 2) Scraper
    print("\n🕷️ 正在调用 Scraper ...")
    scraper_client = ScraperClient(SCRAPER_BASE_URL)
    sc_pois = scraper_client.search(config)
    print(f"✅ 获取到 {len(sc_pois)} 个 POI")

    # 3) 导出 Excel
    print("\n📁 正在生成 Excel ...")
    filepath = create_excel(gm_pois, sc_pois, config, city_name=city_name)
    print(f"✅ 已保存: {filepath}")

    print(f"\n📋 导出统计")
    print(f"   Google Maps API: {len(gm_pois)} 行")
    print(f"   Scraper:         {len(sc_pois)} 行")
    print(f"   文件路径:        {os.path.abspath(filepath)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
